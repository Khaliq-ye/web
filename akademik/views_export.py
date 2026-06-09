# ============================================================
# File: akademik/views_export.py
# PATCH: Perbaikan data Wali Murid di semua section
#   - _build_identitas : tampilkan nama wali dari FK siswa.wali
#   - _build_ttd       : nama wali tetap dari kelas.wali_kelas (guru),
#                        tapi baris "Nama Orang Tua" sudah benar
#   - export_nilai_excel: identitas di semua sheet pakai nama wali
#   - export_semua_siswa_excel: kolom wali murid ditambahkan
#   - select_related   : tambah 'wali' di semua query get siswa
# ============================================================
# ============================================================
# File: akademik/views_export.py
# PATCH: Perbaikan data Wali Murid di semua section
#   - _build_identitas : tampilkan nama wali dari FK siswa.wali
#   - _build_ttd       : nama wali tetap dari kelas.wali_kelas (guru),
#                        tapi baris "Nama Orang Tua" sudah benar
#   - export_nilai_excel: identitas di semua sheet pakai nama wali
#   - export_semua_siswa_excel: kolom wali murid ditambahkan
#   - select_related   : tambah 'wali' di semua query get siswa
# PATCH v2: Tambah foto siswa di identitas rapor PDF
#   - _build_identitas : foto siswa dari DB ditampilkan di kolom kanan
#                        dengan border navy, fallback ke placeholder SVG
# ============================================================
import os
from io import BytesIO
from datetime import date

from django.http import HttpResponse, Http404
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.conf import settings

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable, Image,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Rect

from core.models import (
    Siswa,
    KepalaSekolah as CoreKepsek,
    KeuanganSiswa,
)
from akademik.models import NilaiRapor, NilaiPerilaku, Presensi, JadwalPelajaran
from web_publik.models import KontakSekolah


# ── Palet warna FORMAL ──────────────────────────────────────
NAVY       = colors.HexColor('#1e3a8a')
NAVY_LIGHT = colors.HexColor('#1e40af')
BLUE_LINE  = colors.HexColor('#3b82f6')
ROW_ALT    = colors.HexColor('#f8fafc')
ROW_WHITE  = colors.white
HEADER_BG  = colors.HexColor('#1e3a8a')
SUB_BG     = colors.HexColor('#eff6ff')
GRID_LINE  = colors.HexColor('#d1d5db')
BOX_LINE   = colors.HexColor('#9ca3af')
PASS_CLR   = colors.HexColor('#15803d')
FAIL_CLR   = colors.HexColor('#b91c1c')
WARN_CLR   = colors.HexColor('#92400e')
TEXT_DARK  = colors.HexColor('#111827')
TEXT_GREY  = colors.HexColor('#6b7280')
WHITE      = colors.white
GOLD_LINE  = colors.HexColor('#ca8a04')

PAGE_W, PAGE_H = A4
MARGIN_LR = 2.0 * cm
MARGIN_TB = 1.8 * cm
W = PAGE_W - 2 * MARGIN_LR


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def _hitung_nilai_perilaku(siswa):
    SKOR = {'Baik': 100, 'Cukup': 70, 'Kurang': 30}
    qs = NilaiPerilaku.objects.filter(siswa=siswa).values_list('kategori', flat=True)
    if not qs.exists():
        return 0
    return int(round(sum(SKOR.get(k, 0) for k in qs) / qs.count()))


def _hitung_nilai_akhir(nilai_tugas, nilai_uts, nilai_uas, nilai_perilaku):
    return int(
        (nilai_tugas      * 0.35)
        + (nilai_uts      * 0.25)
        + (nilai_uas      * 0.25)
        + (nilai_perilaku * 0.15)
    )


def _predikat(nilai):
    if nilai >= 90: return 'A'
    if nilai >= 80: return 'B'
    if nilai >= 70: return 'C'
    if nilai >= 60: return 'D'
    return 'E'


def _ket(nilai):
    return 'Tuntas' if nilai >= 70 else 'Belum Tuntas'


def _label_perilaku(skor):
    if skor >= 85: return 'Sangat Baik'
    if skor >= 70: return 'Baik'
    if skor >= 50: return 'Cukup'
    return 'Perlu Peningkatan'


# ── PATCH: Ambil nama wali murid dari FK siswa.wali ──────────
def _nama_wali(siswa):
    """
    Prioritas:
      1. siswa.wali.nama_lengkap  (FK ke WaliMurid — data resmi)
    """
    if siswa.wali and getattr(siswa.wali, 'nama_lengkap', None):
        return siswa.wali.nama_lengkap
    return '-'


# ── PATCH: Ambil no HP wali murid ───────────────────────────
def _no_hp_wali(siswa):
    if siswa.wali and getattr(siswa.wali, 'no_hp', None):
        return siswa.wali.no_hp
    return '-'


# ── PATCH v2: Ambil path foto siswa dari field foto ─────────
def _get_foto_siswa_path(siswa):
    """
    Mengembalikan path absolut foto siswa jika tersedia dan file ada di disk.
    Returns None jika tidak ada foto.
    """
    try:
        if not siswa.foto:
            return None
        # Field ImageField/FileField memiliki atribut .path
        foto_path = siswa.foto.path
        if os.path.exists(foto_path):
            return foto_path
    except (AttributeError, ValueError, NotImplementedError):
        # ValueError jika storage tidak support .path (e.g. remote storage)
        # NotImplementedError pada beberapa custom storage backends
        pass
    return None


def ps(name, size=9, bold=False, color=TEXT_DARK, align=TA_LEFT,
       leading=None, space_before=0, space_after=0):
    fn = 'Helvetica-Bold' if bold else 'Helvetica'
    kw = dict(
        fontSize=size, fontName=fn, textColor=color,
        alignment=align, spaceAfter=space_after, spaceBefore=space_before,
    )
    if leading:
        kw['leading'] = leading
    return ParagraphStyle(name, **kw)


def _get_logo_path():
    candidates = [
        os.path.join(settings.BASE_DIR, 'web_publik', 'static', 'images', 'logo-profil-sekolah.png'),
        os.path.join(settings.BASE_DIR, 'static', 'images', 'logo-profil-sekolah.png'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'images', 'logo-profil-sekolah.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'logo.png'),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def _mini_bar(nilai, width=2.6 * cm, height=0.28 * cm):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height,
               fillColor=colors.HexColor('#e5e7eb'), strokeColor=None))
    bar_w     = width * (nilai / 100)
    bar_color = (PASS_CLR if nilai >= 70 else WARN_CLR if nilai >= 60 else FAIL_CLR)
    d.add(Rect(0, 0, bar_w, height, fillColor=bar_color, strokeColor=None))
    return d


def _hitung_ranking(siswa):
    if not siswa.kelas:
        return None, 0
    semua = Siswa.objects.filter(kelas=siswa.kelas)
    rata_list = []
    for s in semua:
        prl = _hitung_nilai_perilaku(s)
        qs  = NilaiRapor.objects.filter(siswa=s)
        rata = (
            sum(_hitung_nilai_akhir(n.nilai_tugas, n.nilai_uts, n.nilai_uas, prl) for n in qs) / qs.count()
            if qs.exists() else 0
        )
        rata_list.append((s.pk, rata))
    rata_list.sort(key=lambda x: x[1], reverse=True)
    total = len(rata_list)
    for rank, (pk, _) in enumerate(rata_list, 1):
        if pk == siswa.pk:
            return rank, total
    return None, total


def _get_kepsek():
    return CoreKepsek.objects.select_related('user').first()


def _cek_akses(user, siswa_id):
    is_siswa_sendiri = (
        hasattr(user, 'siswa_profile') and user.siswa_profile.pk == siswa_id
    )
    return (
        user.is_staff
        or user.is_superuser
        or hasattr(user, 'kepsek_core_user')
        or hasattr(user, 'guru_profile')
        or is_siswa_sendiri
    )


# ============================================================
# SECTION: HEADER DOKUMEN
# ============================================================
def _build_header(story, kontak, logo_path, nama_sekolah_override=None):
    nama_sekolah = nama_sekolah_override or 'SMA NEGERI 5 KOTA X'
    alamat = kontak.alamat  if kontak and kontak.alamat  else 'Jl. Pendidikan No. 123, Lombok Timur, NTB'
    telp   = kontak.telepon if kontak and kontak.telepon else '(0376) 23456'
    email  = kontak.email   if kontak and kontak.email   else 'info@siakadpro.sch.id'

    LOGO_SIZE = 1.8 * cm

    info_paragraphs = [
        Paragraph(nama_sekolah,
                  ps('hs', 14, bold=True, color=NAVY, align=TA_LEFT)),
        Paragraph(alamat,
                  ps('ha', 8.5, color=TEXT_GREY, leading=12, align=TA_LEFT)),
        Paragraph(f'Telp. {telp}   Email: {email}',
                  ps('hc', 8.5, color=TEXT_GREY, align=TA_LEFT)),
    ]

    if logo_path:
        logo_img   = Image(logo_path, width=LOGO_SIZE, height=LOGO_SIZE)
        hdr_data   = [[logo_img, info_paragraphs]]
        hdr_widths = [LOGO_SIZE + 0.4 * cm, W - LOGO_SIZE - 0.4 * cm]
    else:
        hdr_data   = [['', info_paragraphs]]
        hdr_widths = [0, W]

    hdr_tbl = Table(hdr_data, colWidths=hdr_widths)
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 5))

    story.append(HRFlowable(width=W, thickness=2.5, color=NAVY,      spaceAfter=1))
    story.append(HRFlowable(width=W, thickness=0.8, color=GOLD_LINE, spaceAfter=6))

    banner = Table(
        [[Paragraph('LAPORAN HASIL STUDI (RAPOR)',
                    ps('bn', 12, bold=True, color=WHITE, align=TA_CENTER))]],
        colWidths=[W]
    )
    banner.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(banner)

    sub = Table(
        [[
            Paragraph('Tahun Akademik : 2025 / 2026',
                      ps('s1', 9, color=NAVY, align=TA_LEFT)),
            Paragraph('Semester : Ganjil',
                      ps('s2', 9, color=NAVY, align=TA_CENTER)),
            Paragraph('Tanggal Cetak : ' + date.today().strftime('%d %B %Y'),
                      ps('s3', 9, color=NAVY, align=TA_RIGHT)),
        ]],
        colWidths=[W / 3, W / 3, W / 3]
    )
    sub.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), SUB_BG),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
        ('BOX',           (0, 0), (-1, -1), 0.5, GRID_LINE),
    ]))
    story.append(sub)
    story.append(Spacer(1, 10))


# ============================================================
# SECTION: IDENTITAS SISWA + RANKING + FOTO SISWA
#
# PATCH v2: Foto siswa ditampilkan di kolom paling kanan
#   Layout baru:
#     [Header "A. IDENTITAS SISWA"                         ]
#     [ Tabel 6 kolom biodata (kiri) | Foto siswa (kanan)  ]
#
#   - Foto diambil dari siswa.foto (ImageField di model Siswa)
#   - Ukuran foto: 2.8 cm x 3.5 cm (proporsi 4:5 / pas foto resmi)
#   - Border: navy 1.5pt, background abu terang jika tidak ada foto
#   - Foto di-SPAN 4 baris agar sejajar dengan seluruh biodata
#   - Fallback: kotak placeholder bertuliskan "Foto\nSiswa" jika
#     file tidak ada atau belum diupload
# ============================================================
def _build_identitas(story, siswa, ranking, total_kelas, nilai_perilaku_global):
    nama_kelas = siswa.kelas.nama_kelas if siswa.kelas else '-'
    tgl_lahir  = siswa.tanggal_lahir.strftime('%d %B %Y') if siswa.tanggal_lahir else '-'
    ttl        = (siswa.tempat_lahir or '-') + ', ' + tgl_lahir

    # Wali kelas (guru) dari relasi kelas
    wali_kelas_nama = '-'
    if siswa.kelas and siswa.kelas.wali_kelas:
        wali_kelas_nama = siswa.kelas.wali_kelas.nama_lengkap

    # ── Nama & no HP wali murid dari FK ─────────────────────
    nama_wali_murid = _nama_wali(siswa)
    no_hp_wali      = _no_hp_wali(siswa)
    if no_hp_wali != '-':
        nama_wali_display = f'{nama_wali_murid}  ({no_hp_wali})'
    else:
        nama_wali_display = nama_wali_murid

    label_prl  = _label_perilaku(nilai_perilaku_global)
    warna_prl  = PASS_CLR if nilai_perilaku_global >= 70 else WARN_CLR if nilai_perilaku_global >= 50 else FAIL_CLR
    rank_txt   = f'Ke-{ranking} dari {total_kelas} siswa' if ranking else 'Belum dapat dihitung'

    # ── Header section ───────────────────────────────────────
    hdr = Table(
        [[Paragraph('A.  IDENTITAS SISWA',
                    ps('hid', 9, bold=True, color=WHITE))]],
        colWidths=[W]
    )
    hdr.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(hdr)

    # ── Dimensi foto pas foto resmi (3R: 2.8cm x 3.5cm) ─────
    FOTO_W = 2.8 * cm
    FOTO_H = 3.5 * cm

    # ── Coba muat foto siswa dari storage ───────────────────
    foto_path = _get_foto_siswa_path(siswa)

    if foto_path:
        # Foto tersedia — buat Image dengan dimensi pas foto
        foto_widget = Image(
            foto_path,
            width=FOTO_W,
            height=FOTO_H,
        )
    else:
        # Placeholder: Drawing kotak abu-abu dengan teks "FOTO SISWA"
        from reportlab.graphics.shapes import String
        foto_widget = Drawing(FOTO_W, FOTO_H)
        # Background kotak
        foto_widget.add(Rect(
            0, 0, FOTO_W, FOTO_H,
            fillColor=colors.HexColor('#e8edf5'),
            strokeColor=NAVY,
            strokeWidth=1,
        ))
        # Ikon orang sederhana (lingkaran kepala + trapesoid badan)
        from reportlab.graphics.shapes import Circle, Polygon
        cx = FOTO_W / 2
        # Kepala
        foto_widget.add(Circle(
            cx, FOTO_H * 0.62, FOTO_H * 0.14,
            fillColor=colors.HexColor('#9ca3af'),
            strokeColor=None,
        ))
        # Badan
        foto_widget.add(Polygon(
            [
                cx - FOTO_W * 0.22, FOTO_H * 0.12,
                cx + FOTO_W * 0.22, FOTO_H * 0.12,
                cx + FOTO_W * 0.32, FOTO_H * 0.40,
                cx - FOTO_W * 0.32, FOTO_H * 0.40,
            ],
            fillColor=colors.HexColor('#9ca3af'),
            strokeColor=None,
        ))
        # Label teks
        foto_widget.add(String(
            cx, FOTO_H * 0.05,
            'FOTO SISWA',
            fontSize=5.5,
            fillColor=colors.HexColor('#6b7280'),
            textAnchor='middle',
            fontName='Helvetica',
        ))

    # ── Styles tabel biodata ─────────────────────────────────
    lbl = ps('lbl', 8.5, bold=False, color=TEXT_GREY)
    bld = ps('bld', 9,   bold=True,  color=TEXT_DARK)
    val = ps('val', 9,   bold=False, color=TEXT_DARK)

    # Lebar kolom biodata (6 kolom) — sisakan ruang untuk foto di kolom ke-7
    FOTO_COL_W = FOTO_W + 0.6 * cm   # lebar kolom foto + padding
    BIO_W      = W - FOTO_COL_W       # total lebar area biodata

    C1 = 3.5 * cm    # label kiri
    C2 = 0.4 * cm    # titik dua kiri
    C3 = BIO_W / 2 - C1 - C2 - 0.2 * cm   # nilai kiri
    C4 = 3.0 * cm    # label kanan
    C5 = 0.4 * cm    # titik dua kanan
    C6 = BIO_W - C1 - C2 - C3 - C4 - C5   # nilai kanan

    # ── 4 baris biodata + kolom foto di-SPAN ────────────────
    # Setiap baris: 6 kolom biodata + 1 kolom foto
    # Kolom foto (index 6) di-SPAN dari baris 0 s/d 3
    rows = [
        # Baris 0: Nama + NISN
        [
            Paragraph('Nama Lengkap',     lbl), Paragraph(':', lbl),
            Paragraph(siswa.nama_lengkap,
                      ps('nv', 9, bold=True, color=TEXT_DARK)),
            Paragraph('NISN',             lbl), Paragraph(':', lbl),
            Paragraph(siswa.nisn,          bld),
            foto_widget,   # ← foto siswa, di-SPAN 4 baris
        ],
        # Baris 1: TTL + Kelas
        [
            Paragraph('Tempat, Tgl Lahir', lbl), Paragraph(':', lbl),
            Paragraph(ttl, val),
            Paragraph('Kelas',             lbl), Paragraph(':', lbl),
            Paragraph(nama_kelas,           bld),
            '',   # placeholder — di-SPAN dari baris 0
        ],
        # Baris 2: Wali Murid + Wali Kelas
        [
            Paragraph('Orang Tua / Wali', lbl), Paragraph(':', lbl),
            Paragraph(nama_wali_display,
                      ps('wv', 9, bold=False, color=TEXT_DARK)),
            Paragraph('Wali Kelas',       lbl), Paragraph(':', lbl),
            Paragraph(wali_kelas_nama,     val),
            '',   # placeholder — di-SPAN dari baris 0
        ],
        # Baris 3: Nilai Perilaku + Peringkat
        [
            Paragraph('Nilai Perilaku',   lbl), Paragraph(':', lbl),
            Paragraph(
                f'{nilai_perilaku_global} / 100  ({label_prl})',
                ps('pv', 9, bold=True, color=warna_prl)
            ),
            Paragraph('Peringkat Kelas',  lbl), Paragraph(':', lbl),
            Paragraph(rank_txt,
                      ps('rv', 9, bold=True, color=NAVY)),
            '',   # placeholder — di-SPAN dari baris 0
        ],
    ]

    id_tbl = Table(
        rows,
        colWidths=[C1, C2, C3, C4, C5, C6, FOTO_COL_W],
    )
    id_tbl.setStyle(TableStyle([
        # ── Warna baris bergantian ──────────────────────────
        ('ROWBACKGROUNDS', (0, 0), (5, -1), [ROW_WHITE, ROW_ALT, ROW_WHITE, ROW_ALT]),

        # ── Grid & box biodata (6 kolom pertama) ────────────
        ('GRID',           (0, 0), (5, -1), 0.4, GRID_LINE),
        ('BOX',            (0, 0), (5, -1), 0.8, BOX_LINE),

        # ── Padding seluruh tabel ───────────────────────────
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ('LEFTPADDING',    (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',   (0, 0), (-1, -1), 4),
        ('VALIGN',         (0, 0), (5, -1),  'MIDDLE'),

        # ── Kolom foto (index 6): SPAN 4 baris ──────────────
        ('SPAN',           (6, 0), (6, 3)),

        # ── Styling khusus kolom foto ────────────────────────
        ('VALIGN',         (6, 0), (6, 3),  'MIDDLE'),
        ('ALIGN',          (6, 0), (6, 3),  'CENTER'),
        ('LEFTPADDING',    (6, 0), (6, 3),  6),
        ('RIGHTPADDING',   (6, 0), (6, 3),  6),
        ('TOPPADDING',     (6, 0), (6, 3),  6),
        ('BOTTOMPADDING',  (6, 0), (6, 3),  6),

        # ── Border kolom foto: navy tegas ───────────────────
        ('BOX',            (6, 0), (6, 3),  1.5, NAVY),
        ('BACKGROUND',     (6, 0), (6, 3),  colors.HexColor('#f0f4ff')),
    ]))

    story.append(id_tbl)
    story.append(Spacer(1, 10))


# ============================================================
# SECTION: TABEL NILAI RAPOR  (tidak berubah)
# ============================================================
def _build_nilai(story, nilai_data, rata_rata_rapor, total_mapel):
    hdr = Table(
        [[Paragraph('B.  REKAP NILAI AKADEMIK',
                    ps('hnil', 9, bold=True, color=WHITE))]],
        colWidths=[W]
    )
    hdr.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(hdr)
    story.append(Paragraph(
        'Nilai Akhir = Tugas (35%) + UTS (25%) + UAS (25%) + Perilaku (15%)   |   KKM : 70',
        ps('frm', 8, color=TEXT_GREY, space_before=4, space_after=4)
    ))

    th  = ps('th',  8.5, bold=True,  color=WHITE,     align=TA_CENTER)
    tdc = ps('tdc', 8.5, bold=False, color=TEXT_DARK, align=TA_CENTER)
    tdl = ps('tdl', 8.5, bold=False, color=TEXT_DARK, align=TA_LEFT)

    NW = [0.55*cm, 4.5*cm, 1.3*cm, 1.3*cm, 1.3*cm, 1.4*cm, 1.4*cm, 2.6*cm, 1.0*cm]
    NW.append(W - sum(NW))

    rows = [[
        Paragraph('No',             th),
        Paragraph('Mata Pelajaran', th),
        Paragraph('Tugas\n(35%)',   th),
        Paragraph('UTS\n(25%)',     th),
        Paragraph('UAS\n(25%)',     th),
        Paragraph('Perilaku\n(15%)',th),
        Paragraph('Nilai\nAkhir',   th),
        Paragraph('Grafik',         th),
        Paragraph('Grade',          th),
        Paragraph('Keterangan',     th),
    ]]

    for i, v in enumerate(nilai_data, 1):
        pred  = _predikat(v['akhir'])
        ket   = _ket(v['akhir'])
        warna = PASS_CLR if v['akhir'] >= 70 else FAIL_CLR
        rows.append([
            Paragraph(str(i),            tdc),
            Paragraph(v['mapel'],        tdl),
            Paragraph(str(v['tugas']),   tdc),
            Paragraph(str(v['uts']),     tdc),
            Paragraph(str(v['uas']),     tdc),
            Paragraph(str(v['perilaku']),tdc),
            Paragraph(str(v['akhir']),   ps(f'na{i}', 9, bold=True, color=warna, align=TA_CENTER)),
            _mini_bar(v['akhir']),
            Paragraph(pred,              tdc),
            Paragraph(ket,               ps(f'kt{i}', 8.5, color=warna, align=TA_CENTER)),
        ])

    if total_mapel > 0 and rata_rata_rapor is not None:
        warna_avg = PASS_CLR if rata_rata_rapor >= 70 else FAIL_CLR
        rows.append([
            Paragraph('', tdc),
            Paragraph('RATA-RATA SELURUH MATA PELAJARAN',
                      ps('avgl', 9, bold=True, color=NAVY)),
            Paragraph('', tdc), Paragraph('', tdc),
            Paragraph('', tdc), Paragraph('', tdc),
            Paragraph(str(rata_rata_rapor),
                      ps('avgv', 10, bold=True, color=warna_avg, align=TA_CENTER)),
            Paragraph('', tdc),
            Paragraph(_predikat(int(rata_rata_rapor)), tdc),
            Paragraph(_ket(int(rata_rata_rapor)),
                      ps('avgkt', 9, color=warna_avg, align=TA_CENTER)),
        ])

    n_tbl = Table(rows, colWidths=NW, repeatRows=1)
    n_tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0),  (-1, 0),  HEADER_BG),
        ('ROWBACKGROUNDS', (0, 1),  (-1, -2), [ROW_WHITE, ROW_ALT]),
        ('BACKGROUND',     (0, -1), (-1, -1), SUB_BG),
        ('GRID',           (0, 0),  (-1, -1), 0.4, GRID_LINE),
        ('BOX',            (0, 0),  (-1, -1), 1,   BOX_LINE),
        ('TOPPADDING',     (0, 0),  (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0),  (-1, -1), 4),
        ('LEFTPADDING',    (0, 0),  (-1, -1), 3),
        ('RIGHTPADDING',   (0, 0),  (-1, -1), 3),
        ('VALIGN',         (0, 0),  (-1, -1), 'MIDDLE'),
        ('ALIGN',          (0, 0),  (-1, -1), 'CENTER'),
        ('ALIGN',          (1, 0),  (1, -1),  'LEFT'),
        ('LINEBELOW',      (0, 0),  (-1, 0),  1.2, BLUE_LINE),
    ]))
    story.append(n_tbl)
    story.append(Spacer(1, 10))


# ============================================================
# SECTION: REKAP KEHADIRAN  (tidak berubah)
# ============================================================
def _build_kehadiran(story, hadir, izin, sakit, alpa, total, pct):
    hdr = Table(
        [[Paragraph('C.  REKAP KEHADIRAN',
                    ps('hkh', 9, bold=True, color=WHITE))]],
        colWidths=[W]
    )
    hdr.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(hdr)

    pct_color = PASS_CLR if pct >= 75 else FAIL_CLR
    th = ps('kh', 8.5, bold=True, color=WHITE, align=TA_CENTER)
    tv = ps('kv', 11,  bold=True, color=NAVY,  align=TA_CENTER)
    ts = ps('ks', 8,   color=TEXT_GREY,         align=TA_CENTER)
    tp = ps('kp', 12,  bold=True, color=pct_color, align=TA_CENTER)

    att = [
        [Paragraph('Hadir', th), Paragraph('Izin', th),
         Paragraph('Sakit', th), Paragraph('Alpa', th),
         Paragraph('Total', th), Paragraph('% Kehadiran', th)],
        [Paragraph(str(hadir), tv), Paragraph(str(izin), tv),
         Paragraph(str(sakit), tv), Paragraph(str(alpa), tv),
         Paragraph(str(total), tv), Paragraph(f'{pct}%', tp)],
        [Paragraph('hari', ts), Paragraph('hari', ts),
         Paragraph('hari', ts), Paragraph('hari', ts),
         Paragraph('hari', ts), Paragraph('dari total hari', ts)],
    ]

    a_tbl = Table(att, colWidths=[W / 6] * 6)
    a_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  HEADER_BG),
        ('BACKGROUND',    (0, 1), (-1, 2),  ROW_WHITE),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.2, BLUE_LINE),
        ('GRID',          (0, 0), (-1, -1), 0.4, GRID_LINE),
        ('BOX',           (0, 0), (-1, -1), 1,   BOX_LINE),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(a_tbl)
    story.append(Spacer(1, 10))


# ============================================================
# SECTION: PENILAIAN PERILAKU  (tidak berubah)
# ============================================================
def _build_perilaku(story, skor, total_penilaian, baik, cukup, kurang):
    warna = PASS_CLR if skor >= 70 else WARN_CLR if skor >= 50 else FAIL_CLR
    label = _label_perilaku(skor)

    hdr = Table(
        [[Paragraph('D.  PENILAIAN PERILAKU DAN SIKAP',
                    ps('hprl', 9, bold=True, color=WHITE))]],
        colWidths=[W]
    )
    hdr.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(hdr)

    C1, C2 = 3.8 * cm, 0.5 * cm
    C3     = W / 2 - C1 - C2 - 0.2 * cm
    C4, C5 = 3.2 * cm, 0.5 * cm
    C6     = W - C1 - C2 - C3 - C4 - C5
    lbl = ps('pl', 8.5, color=TEXT_GREY)
    val = ps('pv', 9,   color=TEXT_DARK)

    rows = [
        [Paragraph('Skor Perilaku',    lbl), Paragraph(':', lbl),
         Paragraph(f'{skor} / 100',    ps('ps1', 10, bold=True, color=warna)),
         Paragraph('Kategori',         lbl), Paragraph(':', lbl),
         Paragraph(label,              ps('ps2', 10, bold=True, color=warna))],

        [Paragraph('Total Penilaian',  lbl), Paragraph(':', lbl),
         Paragraph(f'{total_penilaian} penilaian dari guru', val),
         Paragraph('Kontribusi Rapor', lbl), Paragraph(':', lbl),
         Paragraph('15% dari Nilai Akhir',
                   ps('ps3', 9, bold=True, color=NAVY))],

        [Paragraph('Distribusi',       lbl), Paragraph(':', lbl),
         Paragraph(f'Baik ({baik})     Cukup ({cukup})     Kurang ({kurang})', val),
         Paragraph('Skala',            lbl), Paragraph(':', lbl),
         Paragraph('Baik = 100   Cukup = 70   Kurang = 30',
                   ps('ps4', 8.5, color=TEXT_GREY))],
    ]

    prl_tbl = Table(rows, colWidths=[C1, C2, C3, C4, C5, C6])
    prl_tbl.setStyle(TableStyle([
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [ROW_WHITE, ROW_ALT, ROW_WHITE]),
        ('GRID',           (0, 0), (-1, -1), 0.4, GRID_LINE),
        ('BOX',            (0, 0), (-1, -1), 0.8, BOX_LINE),
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ('LEFTPADDING',    (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',   (0, 0), (-1, -1), 4),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(prl_tbl)
    story.append(Spacer(1, 10))


# ============================================================
# SECTION: CATATAN DAN KESIMPULAN  (tidak berubah)
# ============================================================
def _build_catatan(story, rata_rapor, ranking, total_kelas):
    if rata_rapor is None:
        kesimpulan = 'Belum ada data nilai pada semester ini.'
    elif rata_rapor >= 85:
        kesimpulan = (
            f'Siswa menunjukkan prestasi akademik yang sangat memuaskan '
            f'dengan rata-rata nilai {rata_rapor}. '
            'Pertahankan semangat belajar dan terus tingkatkan prestasi.'
        )
    elif rata_rapor >= 70:
        kesimpulan = (
            f'Siswa telah memenuhi standar ketuntasan belajar dengan '
            f'rata-rata nilai {rata_rapor}. '
            'Tingkatkan fokus pada mata pelajaran yang masih di bawah KKM.'
        )
    else:
        kesimpulan = (
            f'Siswa perlu meningkatkan prestasi belajar. '
            f'Rata-rata nilai {rata_rapor} masih di bawah KKM (70). '
            'Diharapkan siswa lebih giat belajar dan aktif berkonsultasi dengan guru.'
        )

    rank_txt  = f'Ke-{ranking} dari {total_kelas} siswa' if ranking else 'Belum dapat dihitung'
    warna_avg = PASS_CLR if rata_rapor and rata_rapor >= 70 else FAIL_CLR

    hdr = Table(
        [[Paragraph('E.  CATATAN DAN KESIMPULAN',
                    ps('hcat', 9, bold=True, color=WHITE))]],
        colWidths=[W]
    )
    hdr.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(hdr)

    C1, C2 = 3.8 * cm, 0.5 * cm
    C3     = W / 2 - C1 - C2 - 0.2 * cm
    C4, C5 = 3.2 * cm, 0.5 * cm
    C6     = W - C1 - C2 - C3 - C4 - C5
    lbl = ps('cl', 8.5, color=TEXT_GREY)

    rekomendasi_teks = (
        'Direkomendasikan Naik Kelas'
        if rata_rapor and rata_rapor >= 70
        else 'Perlu Evaluasi Lebih Lanjut'
    )
    rekomendasi_warna = PASS_CLR if rata_rapor and rata_rapor >= 70 else WARN_CLR

    cat_rows = [
        [Paragraph('Rata-rata Nilai',  lbl), Paragraph(':', lbl),
         Paragraph(str(rata_rapor) if rata_rapor else '-',
                   ps('rv', 10, bold=True, color=warna_avg)),
         Paragraph('Peringkat Kelas',  lbl), Paragraph(':', lbl),
         Paragraph(rank_txt, ps('rkv', 9, bold=True, color=NAVY))],

        [Paragraph('Status KKM',       lbl), Paragraph(':', lbl),
         Paragraph('Tuntas' if rata_rapor and rata_rapor >= 70 else 'Perlu Remedial',
                   ps('skv', 9, bold=True, color=warna_avg)),
         Paragraph('Rekomendasi',      lbl), Paragraph(':', lbl),
         Paragraph(rekomendasi_teks,
                   ps('rec', 9, bold=True, color=rekomendasi_warna))],

        [Paragraph('Catatan Wali Kelas', lbl), Paragraph(':', lbl),
         Paragraph(kesimpulan,
                   ps('cat', 8.5, color=TEXT_DARK, leading=13)),
         Paragraph('', lbl), Paragraph('', lbl), Paragraph('', lbl)],
    ]

    cat_tbl = Table(cat_rows, colWidths=[C1, C2, C3, C4, C5, C6])
    cat_tbl.setStyle(TableStyle([
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [ROW_WHITE, ROW_ALT, ROW_WHITE]),
        ('GRID',           (0, 0), (-1, -1), 0.4, GRID_LINE),
        ('BOX',            (0, 0), (-1, -1), 0.8, BOX_LINE),
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ('LEFTPADDING',    (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',   (0, 0), (-1, -1), 4),
        ('VALIGN',         (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN',           (2, 2), (-1, 2)),
        ('VALIGN',         (0, 2), (-1, 2), 'TOP'),
    ]))
    story.append(cat_tbl)
    story.append(Spacer(1, 14))


# ============================================================
# SECTION: TANDA TANGAN
# PATCH: baris "Orang Tua / Wali Siswa" sekarang tampilkan
#        nama_wali dari FK siswa.wali atau fallback
# ============================================================
def _build_ttd(story, siswa, kepsek):
    # Wali kelas (guru)
    wali_kelas_nama = '..............................'
    wali_kelas_nip  = '-'
    if siswa.kelas and siswa.kelas.wali_kelas:
        wali_kelas_nama = siswa.kelas.wali_kelas.nama_lengkap
        wali_kelas_nip  = getattr(siswa.kelas.wali_kelas, 'nip', '-') or '-'

    # Kepala sekolah
    kepsek_nama = kepsek.nama_lengkap if kepsek else '..............................'
    kepsek_nip  = kepsek.nip          if kepsek else '-'

    # ── PATCH: nama wali murid untuk kolom TTD ───────────────
    nama_wali_ttd = _nama_wali(siswa)
    if nama_wali_ttd == '-':
        nama_wali_ttd = '..............................'

    kota_tgl = 'Lombok Timur, ' + date.today().strftime('%d %B %Y')
    tc = ps('tc', 8.5, align=TA_CENTER, color=TEXT_DARK)
    tb = ps('tb', 8.5, bold=True, align=TA_CENTER, color=TEXT_DARK)
    tn = ps('tn', 8,   align=TA_CENTER, color=TEXT_GREY)

    ttd = Table([
        [
            Paragraph(kota_tgl + '<br/><br/><br/><br/><br/>Orang Tua / Wali Siswa', tc),
            Paragraph(kota_tgl + '<br/><br/><br/><br/><br/>Wali Kelas', tc),
            Paragraph(kota_tgl + '<br/><br/><br/><br/><br/>Kepala Sekolah', tc),
        ],
        [
            Paragraph(f'( {nama_wali_ttd} )', tb),
            Paragraph(f'( {wali_kelas_nama} )', tb),
            Paragraph(f'( {kepsek_nama} )', tb),
        ],
        [
            Paragraph('', tn),
            Paragraph(f'NIP. {wali_kelas_nip}', tn),
            Paragraph(f'NIP. {kepsek_nip}', tn),
        ],
    ], colWidths=[W / 3, W / 3, W / 3])

    ttd.setStyle(TableStyle([
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX',           (0, 0), (-1, -1), 0.5, GRID_LINE),
        ('GRID',          (0, 0), (-1, -1), 0.3, colors.HexColor('#f3f4f6')),
    ]))
    story.append(ttd)
    story.append(Spacer(1, 10))

    story.append(HRFlowable(width=W, thickness=0.8, color=GOLD_LINE, spaceAfter=2))
    story.append(HRFlowable(width=W, thickness=2,   color=NAVY,      spaceAfter=4))
    story.append(Paragraph(
        f'Dicetak pada {date.today().strftime("%d %B %Y")}   |   '
        'SMA NEGERI 5 KOTA X — Sistem Informasi Akademik   |   '
        'Dokumen ini sah tanpa stempel basah',
        ps('ftr', 7.5, color=TEXT_GREY, align=TA_CENTER)
    ))


# ============================================================
# VIEW 1 — Generate PDF rapor individual
# PATCH: tambah 'wali' ke select_related
# ============================================================
@login_required
def export_rapor_pdf(request, siswa_id):
    try:
        siswa = Siswa.objects.select_related(
            'kelas', 'kelas__wali_kelas',
            'wali',   # ← PATCH: load FK wali murid sekaligus
            'user',
        ).get(pk=siswa_id)
    except Siswa.DoesNotExist:
        raise Http404("Data siswa tidak ditemukan.")

    if not _cek_akses(request.user, siswa_id):
        raise Http404("Akses ditolak.")

    nilai_list            = NilaiRapor.objects.filter(siswa=siswa).select_related('mata_pelajaran').order_by('mata_pelajaran__nama')
    nilai_perilaku_global = _hitung_nilai_perilaku(siswa)

    nilai_data = []
    for n in nilai_list:
        akhir = _hitung_nilai_akhir(n.nilai_tugas, n.nilai_uts, n.nilai_uas, nilai_perilaku_global)
        nilai_data.append({
            'mapel'   : n.mata_pelajaran.nama,
            'tugas'   : n.nilai_tugas,
            'uts'     : n.nilai_uts,
            'uas'     : n.nilai_uas,
            'perilaku': nilai_perilaku_global,
            'akhir'   : akhir,
        })

    total_mapel     = len(nilai_data)
    rata_rata_rapor = (
        round(sum(v['akhir'] for v in nilai_data) / total_mapel, 1)
        if total_mapel > 0 else None
    )

    presensi_qs = Presensi.objects.filter(siswa=siswa)
    total_absen = presensi_qs.count()
    hadir_count = presensi_qs.filter(status='Hadir').count()
    izin_count  = presensi_qs.filter(status='Izin').count()
    sakit_count = presensi_qs.filter(status='Sakit').count()
    alpa_count  = presensi_qs.filter(status='Alpa').count()
    pct_hadir   = int(hadir_count / total_absen * 100) if total_absen else 0

    qs_prl          = NilaiPerilaku.objects.filter(siswa=siswa)
    total_penilaian = qs_prl.count()
    perilaku_baik   = qs_prl.filter(kategori='Baik').count()
    perilaku_cukup  = qs_prl.filter(kategori='Cukup').count()
    perilaku_kurang = qs_prl.filter(kategori='Kurang').count()

    ranking, total_kelas = _hitung_ranking(siswa)
    kontak    = KontakSekolah.objects.first()
    kepsek    = _get_kepsek()
    logo_path = _get_logo_path()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_LR, rightMargin=MARGIN_LR,
        topMargin=MARGIN_TB,  bottomMargin=MARGIN_TB,
    )
    story = []
    _build_header(story, kontak, logo_path)
    _build_identitas(story, siswa, ranking, total_kelas, nilai_perilaku_global)
    _build_nilai(story, nilai_data, rata_rata_rapor, total_mapel)
    _build_kehadiran(story, hadir_count, izin_count, sakit_count, alpa_count, total_absen, pct_hadir)
    _build_perilaku(story, nilai_perilaku_global, total_penilaian, perilaku_baik, perilaku_cukup, perilaku_kurang)
    _build_catatan(story, rata_rata_rapor, ranking, total_kelas)
    _build_ttd(story, siswa, kepsek)

    doc.build(story)
    buffer.seek(0)

    nama_file   = f'Rapor_{siswa.nama_lengkap.replace(" ", "_")}_2025-2026.pdf'
    force_dl    = request.GET.get('download') == '1'
    disposition = 'attachment' if force_dl else 'inline'

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'{disposition}; filename="{nama_file}"'
    response['X-Frame-Options']         = 'SAMEORIGIN'
    response['Content-Security-Policy'] = "frame-ancestors 'self'"
    return response


# ============================================================
# VIEW 2 — Halaman wrapper viewer rapor  (tidak berubah)
# ============================================================
@login_required
def lihat_rapor_siswa(request, siswa_id):
    siswa = get_object_or_404(
        Siswa.objects.select_related('kelas'), pk=siswa_id
    )
    return render(request, 'akademik/siswa/lihat_rapor.html', {'siswa': siswa})


# ============================================================
# VIEW 3 — Export Excel nilai satu siswa
# PATCH: identitas di semua sheet pakai _nama_wali()
#        dan tambah 'wali' di select_related
# ============================================================

from openpyxl.utils import get_column_letter

@login_required
def export_nilai_excel(request, siswa_id):
    try:
        siswa = Siswa.objects.select_related(
            'kelas', 'kelas__wali_kelas', 'wali',
        ).get(pk=siswa_id)
    except Siswa.DoesNotExist:
        raise Http404("Data siswa tidak ditemukan.")

    if not _cek_akses(request.user, siswa_id):
        raise Http404("Akses ditolak.")

    nilai_list            = NilaiRapor.objects.filter(siswa=siswa).select_related('mata_pelajaran').order_by('mata_pelajaran__nama')
    nilai_perilaku_global = _hitung_nilai_perilaku(siswa)

    presensi_qs = Presensi.objects.filter(siswa=siswa)
    total_absen = presensi_qs.count()
    hadir_count = presensi_qs.filter(status='Hadir').count()
    izin_count  = presensi_qs.filter(status='Izin').count()
    sakit_count = presensi_qs.filter(status='Sakit').count()
    alpa_count  = presensi_qs.filter(status='Alpa').count()
    pct_hadir   = int(hadir_count / total_absen * 100) if total_absen else 0

    total_penilaian = NilaiPerilaku.objects.filter(siswa=siswa).count()
    perilaku_baik   = NilaiPerilaku.objects.filter(siswa=siswa, kategori='Baik').count()
    perilaku_cukup  = NilaiPerilaku.objects.filter(siswa=siswa, kategori='Cukup').count()
    perilaku_kurang = NilaiPerilaku.objects.filter(siswa=siswa, kategori='Kurang').count()

    ranking, total_kelas = _hitung_ranking(siswa)
    kepsek = _get_kepsek()

    # ── Palet warna ──────────────────────────────────────────
    NAVY        = '1E3A8A'
    NAVY_MID    = '1E40AF'
    NAVY_LIGHT  = 'DBEAFE'
    NAVY_XLIGHT = 'EFF6FF'
    GOLD_LIGHT  = 'FEF9C3'
    GREEN       = '15803D'
    GREEN_LIGHT = 'DCFCE7'
    RED         = 'B91C1C'
    RED_LIGHT   = 'FEE2E2'
    ORANGE      = '92400E'
    GREY_DARK   = '374151'
    GREY_MED    = '6B7280'
    GREY_LIGHT  = 'F3F4F6'
    GREY_LINE   = 'D1D5DB'
    WHITE       = 'FFFFFF'
    ROW_ALT     = 'F8FAFC'

    def _thin(color=GREY_LINE):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def _font(bold=False, size=9, color=GREY_DARK, italic=False):
        return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)

    def _align(h='center', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    nama_kelas      = siswa.kelas.nama_kelas if siswa.kelas else '-'
    wali_kelas_nama = siswa.kelas.wali_kelas.nama_lengkap if (siswa.kelas and siswa.kelas.wali_kelas) else '-'
    wali_kelas_nip  = getattr(siswa.kelas.wali_kelas, 'nip', '-') if (siswa.kelas and siswa.kelas.wali_kelas) else '-'
    rank_txt        = f'Ke-{ranking} dari {total_kelas} siswa' if ranking else 'N/A'
    kepsek_nama     = kepsek.nama_lengkap if kepsek else '-'
    kepsek_nip      = kepsek.nip          if kepsek else '-'
    nama_wali       = _nama_wali(siswa)
    no_hp_wali      = _no_hp_wali(siswa)

    # ── Helper identitas: A:B=label | C=':' | D:I=nilai ──────
    def write_id_row(ws, row, label, value, val_color=GREY_DARK, val_bold=False, bg=WHITE):
        ws.merge_cells(f'A{row}:B{row}')
        c_lbl = ws.cell(row=row, column=1, value=label)
        c_lbl.font      = _font(size=9, color=GREY_MED)
        c_lbl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_lbl.fill      = _fill(bg)
        c_lbl.border    = _thin()
        ws.cell(row=row, column=2).fill   = _fill(bg)
        ws.cell(row=row, column=2).border = _thin()

        c_sep = ws.cell(row=row, column=3, value=':')
        c_sep.font      = _font(color=GREY_MED)
        c_sep.alignment = _align('center')
        c_sep.fill      = _fill(bg)
        c_sep.border    = _thin()

        ws.merge_cells(f'D{row}:I{row}')
        c_val = ws.cell(row=row, column=4, value=value)
        c_val.font      = _font(bold=val_bold, color=val_color, size=10 if val_bold else 9)
        c_val.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_val.fill      = _fill(bg)
        c_val.border    = _thin()
        for col in range(5, 10):
            ws.cell(row=row, column=col).fill   = _fill(bg)
            ws.cell(row=row, column=col).border = _thin()
        ws.row_dimensions[row].height = 22

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — NILAI RAPOR
    # ══════════════════════════════════════════════════════════
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Nilai Rapor'
    ws1.sheet_properties.tabColor = NAVY
    ws1.sheet_view.showGridLines  = False

    # A=No(5) | B=Mapel(28) | C=Tugas(10) | D=UTS(10) | E=UAS(10)
    # F=Perilaku(12) | G=NilaiAkhir(12) | H=Grade(8) | I=Keterangan(14)
    for col_ltr, w in [('A',5),('B',28),('C',10),('D',10),('E',10),('F',12),('G',12),('H',8),('I',14)]:
        ws1.column_dimensions[col_ltr].width = w

    # Baris 1: Banner utama
    ws1.merge_cells('A1:I1')
    c = ws1['A1']
    c.value     = 'LAPORAN HASIL STUDI (RAPOR)  —  SMA NEGERI 5 KOTA X'
    c.font      = Font(name='Arial', bold=True, color=WHITE, size=14)
    c.fill      = _fill(NAVY)
    c.alignment = _align('center')
    ws1.row_dimensions[1].height = 36

    # Baris 2: Sub-header
    ws1.merge_cells('A2:I2')
    c = ws1['A2']
    c.value     = (
        'Tahun Akademik 2025 / 2026   |   Semester Ganjil'
        '   |   Tanggal Cetak : ' + date.today().strftime('%d %B %Y')
    )
    c.font      = Font(name='Arial', size=9, color=NAVY, italic=True)
    c.fill      = _fill(NAVY_XLIGHT)
    c.alignment = _align('center')
    ws1.row_dimensions[2].height = 18

    # Baris 3: Formula info
    ws1.merge_cells('A3:I3')
    c = ws1['A3']
    c.value     = 'Nilai Akhir  =  Tugas (35%)  +  UTS (25%)  +  UAS (25%)  +  Perilaku (15%)     |     KKM : 70'
    c.font      = Font(name='Arial', size=8.5, color=GREY_MED, italic=True)
    c.fill      = _fill(GOLD_LIGHT)
    c.alignment = _align('center')
    ws1.row_dimensions[3].height = 16

    # Baris 4: Spacer
    ws1.row_dimensions[4].height = 12

    # Baris 5: Section A header
    ws1.merge_cells('A5:I5')
    c = ws1['A5']
    c.value     = '  \u258c  A.  IDENTITAS SISWA'
    c.font      = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill      = _fill(NAVY_MID)
    c.alignment = _align('left')
    ws1.row_dimensions[5].height = 22

    # Baris 6–13: Data identitas
    id_rows = [
        ('Nama Lengkap',      siswa.nama_lengkap,  NAVY,      True,  NAVY_XLIGHT),
        ('NISN',              siswa.nisn,           GREY_DARK, False, WHITE),
        ('Kelas',             nama_kelas,           GREY_DARK, True,  GREY_LIGHT),
        ('Wali Kelas',        wali_kelas_nama,      GREY_DARK, False, WHITE),
        ('Orang Tua / Wali',  nama_wali,            GREY_DARK, False, GREY_LIGHT),
        ('No. HP Wali',       no_hp_wali,           GREY_DARK, False, WHITE),
        ('Peringkat Kelas',   rank_txt,             GREEN,     True,  GREY_LIGHT),
        ('Kepala Sekolah',    kepsek_nama,          GREY_DARK, False, WHITE),
    ]
    for i, (lbl, val, vc, vb, bg) in enumerate(id_rows, start=6):
        write_id_row(ws1, i, lbl, val, val_color=vc, val_bold=vb, bg=bg)

    # Baris 14: Spacer
    ws1.row_dimensions[14].height = 14

    # Baris 15: Section B header
    ws1.merge_cells('A15:I15')
    c = ws1['A15']
    c.value     = '  \u258c  B.  REKAP NILAI AKADEMIK'
    c.font      = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill      = _fill(NAVY_MID)
    c.alignment = _align('left')
    ws1.row_dimensions[15].height = 22

    # Baris 16: Header tabel nilai
    # A=No | B=Mata Pelajaran | C=Tugas | D=UTS | E=UAS
    # F=Perilaku | G=Nilai Akhir | H=Grade | I=Keterangan
    headers_tbl = [
        'No', 'Mata Pelajaran', 'Tugas\n(35%)', 'UTS\n(25%)',
        'UAS\n(25%)', 'Perilaku\n(15%)', 'Nilai\nAkhir', 'Grade', 'Keterangan'
    ]
    for col, h in enumerate(headers_tbl, start=1):
        c = ws1.cell(row=16, column=col, value=h)
        c.font      = Font(name='Arial', bold=True, color=WHITE, size=9)
        c.fill      = _fill(NAVY)   # semua header kolom warna sama
        c.alignment = _align('center', wrap=True)
        c.border    = _thin(WHITE)
    ws1.row_dimensions[16].height = 32

    # Baris 17+: Data nilai
    DATA_START      = 17
    nilai_akhir_sum = 0
    jumlah_mapel    = 0

    for i, n in enumerate(nilai_list):
        row   = DATA_START + i
        akhir = _hitung_nilai_akhir(n.nilai_tugas, n.nilai_uts, n.nilai_uas, nilai_perilaku_global)
        pred  = _predikat(akhir)
        ket   = _ket(akhir)
        bg_r  = ROW_ALT if i % 2 else WHITE
        warna = GREEN if akhir >= 70 else RED

        cells = [
            i+1, n.mata_pelajaran.nama,
            n.nilai_tugas, n.nilai_uts, n.nilai_uas,
            nilai_perilaku_global, akhir, pred, ket
        ]
        for col, val in enumerate(cells, start=1):
            c = ws1.cell(row=row, column=col, value=val)
            c.border = _thin()
            if col == 7:       # Nilai Akhir
                c.font      = Font(name='Arial', bold=True, size=10, color=warna)
                c.fill      = _fill(GREEN_LIGHT if akhir >= 70 else RED_LIGHT)
                c.alignment = _align('center')
            elif col == 9:     # Keterangan
                c.font      = Font(name='Arial', bold=True, size=9, color=warna)
                c.fill      = _fill(bg_r)
                c.alignment = _align('center')
            elif col == 2:     # Mata Pelajaran
                c.font      = _font()
                c.fill      = _fill(bg_r)
                c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            else:              # No, Tugas, UTS, UAS, Perilaku, Grade
                c.font      = _font()
                c.fill      = _fill(bg_r)
                c.alignment = _align('center')
        ws1.row_dimensions[row].height = 20
        nilai_akhir_sum += akhir
        jumlah_mapel    += 1

    # ── Baris rata-rata — semua sel A-I ter-fill & border ────
    rata_row = DATA_START + jumlah_mapel
    rata_val = round(nilai_akhir_sum / jumlah_mapel, 1) if jumlah_mapel > 0 else 0
    warna_r  = GREEN if rata_val >= 70 else RED

    ws1.merge_cells(f'A{rata_row}:F{rata_row}')
    for col in range(1, 10):
        c = ws1.cell(row=rata_row, column=col)
        c.border = _thin(NAVY)
        if col <= 6:
            c.fill = _fill(NAVY_LIGHT)
        elif col == 7:
            c.fill = _fill(GREEN_LIGHT if rata_val >= 70 else RED_LIGHT)
        else:
            c.fill = _fill(NAVY_LIGHT)

    c_lbl = ws1.cell(row=rata_row, column=1, value='RATA-RATA SELURUH MATA PELAJARAN')
    c_lbl.font      = Font(name='Arial', bold=True, size=10, color=NAVY)
    c_lbl.alignment = _align('center')

    c7 = ws1.cell(row=rata_row, column=7, value=rata_val)
    c7.font      = Font(name='Arial', bold=True, size=12, color=warna_r)
    c7.alignment = _align('center')

    c8 = ws1.cell(row=rata_row, column=8, value=_predikat(int(rata_val)))
    c8.font      = Font(name='Arial', bold=True, size=10, color=warna_r)
    c8.alignment = _align('center')

    c9 = ws1.cell(row=rata_row, column=9, value=_ket(int(rata_val)))
    c9.font      = Font(name='Arial', bold=True, size=10, color=warna_r)
    c9.alignment = _align('center')

    ws1.row_dimensions[rata_row].height     = 26
    ws1.row_dimensions[rata_row + 1].height = 14

    # ── Section C: Rekap Kehadiran ────────────────────────────
    sec_c = rata_row + 2
    ws1.merge_cells(f'A{sec_c}:I{sec_c}')
    c = ws1[f'A{sec_c}']
    c.value = '  \u258c  C.  REKAP KEHADIRAN'
    c.font  = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill  = _fill(NAVY_MID)
    c.alignment = _align('left')
    ws1.row_dimensions[sec_c].height = 22

    hdr_c    = sec_c + 1
    att_cols = ['Hadir', 'Izin', 'Sakit', 'Alpa', 'Total Hari', '', '', '', '% Kehadiran']
    att_vals = [hadir_count, izin_count, sakit_count, alpa_count, total_absen, '', '', '', f'{pct_hadir}%']
    att_clrs = [GREEN, ORANGE, ORANGE, RED, NAVY, GREY_DARK, GREY_DARK, GREY_DARK, GREEN if pct_hadir >= 75 else RED]

    # FIX: semua header kehadiran fill NAVY agar menyambung rapi
    for col, h in enumerate(att_cols, start=1):
        c = ws1.cell(row=hdr_c, column=col, value=h)
        c.font      = Font(name='Arial', bold=True, size=9, color=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _align('center')
        c.border    = _thin()
    ws1.row_dimensions[hdr_c].height = 22

    val_c = hdr_c + 1
    for col, (val, clr) in enumerate(zip(att_vals, att_clrs), start=1):
        c = ws1.cell(row=val_c, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, size=13 if col in (1,5,9) else 11, color=clr)
        c.fill      = _fill(GREY_LIGHT)
        c.alignment = _align('center')
        c.border    = _thin()
    ws1.row_dimensions[val_c].height = 28

    sub_c   = val_c + 1
    sub_lbl = ['hari','hari','hari','hari','hari','','','','dari total hari']
    for col, lbl in enumerate(sub_lbl, start=1):
        c = ws1.cell(row=sub_c, column=col, value=lbl)
        c.font      = Font(name='Arial', size=8, color=GREY_MED, italic=True)
        c.fill      = _fill(WHITE)
        c.alignment = _align('center')
        c.border    = _thin()
    ws1.row_dimensions[sub_c].height         = 16
    ws1.row_dimensions[sub_c + 1].height     = 14

    # ── Section D: Tanda Tangan ───────────────────────────────
    ttd_row = sub_c + 2
    ws1.merge_cells(f'A{ttd_row}:I{ttd_row}')
    c = ws1[f'A{ttd_row}']
    c.value = '  \u258c  D.  TANDA TANGAN DAN PENGESAHAN'
    c.font  = Font(name='Arial', bold=True, color=WHITE, size=10)
    c.fill  = _fill(NAVY_MID)
    c.alignment = _align('left')
    ws1.row_dimensions[ttd_row].height = 22

    nama_wali_ttd = nama_wali if nama_wali != '-' else '..............................'
    ttd_data  = [
        ('Orang Tua / Wali Siswa', nama_wali_ttd,  ''),
        ('Wali Kelas',             wali_kelas_nama, f'NIP. {wali_kelas_nip}'),
        ('Kepala Sekolah',         kepsek_nama,     f'NIP. {kepsek_nip}'),
    ]
    # Panel 1=A:C | Panel 2=D:F | Panel 3=G:I
    ttd_spans = [(1, 3), (4, 6), (7, 9)]
    kota_tgl  = 'Lombok Timur, ' + date.today().strftime('%d %B %Y')

    for (col_s, col_e), (jabatan, nama, nip) in zip(ttd_spans, ttd_data):
        cs = get_column_letter(col_s)
        ce = get_column_letter(col_e)

        # r1: kota & tanggal
        r1 = ttd_row + 1
        ws1.merge_cells(f'{cs}{r1}:{ce}{r1}')
        c = ws1.cell(row=r1, column=col_s, value=kota_tgl)
        c.font      = _font(size=9, color=GREY_MED, italic=True)
        c.alignment = _align('center')
        for col in range(col_s, col_e + 1):
            ws1.cell(row=r1, column=col).fill   = _fill(WHITE)
            ws1.cell(row=r1, column=col).border = _thin()

        # r2: jabatan
        r2 = ttd_row + 2
        ws1.merge_cells(f'{cs}{r2}:{ce}{r2}')
        c = ws1.cell(row=r2, column=col_s, value=jabatan)
        c.font      = _font(size=9, color=GREY_MED)
        c.alignment = _align('center')
        for col in range(col_s, col_e + 1):
            ws1.cell(row=r2, column=col).fill   = _fill(GREY_LIGHT)
            ws1.cell(row=r2, column=col).border = _thin()

        # r3-r5: ruang tanda tangan kosong
        for rr in range(ttd_row + 3, ttd_row + 6):
            ws1.merge_cells(f'{cs}{rr}:{ce}{rr}')
            c = ws1.cell(row=rr, column=col_s, value='')
            for col in range(col_s, col_e + 1):
                ws1.cell(row=rr, column=col).fill   = _fill(WHITE)
                ws1.cell(row=rr, column=col).border = _thin()
            ws1.row_dimensions[rr].height = 18

        # r6: nama
        r6 = ttd_row + 6
        ws1.merge_cells(f'{cs}{r6}:{ce}{r6}')
        c = ws1.cell(row=r6, column=col_s, value=f'( {nama} )')
        c.font      = Font(name='Arial', bold=True, size=9, color=NAVY, underline='single')
        c.alignment = _align('center')
        for col in range(col_s, col_e + 1):
            ws1.cell(row=r6, column=col).fill   = _fill(NAVY_XLIGHT)
            ws1.cell(row=r6, column=col).border = _thin()

        # r7: NIP
        r7 = ttd_row + 7
        ws1.merge_cells(f'{cs}{r7}:{ce}{r7}')
        c = ws1.cell(row=r7, column=col_s, value=nip or '')
        c.font      = _font(size=8, color=GREY_MED)
        c.alignment = _align('center')
        for col in range(col_s, col_e + 1):
            ws1.cell(row=r7, column=col).fill   = _fill(WHITE)
            ws1.cell(row=r7, column=col).border = _thin()
        ws1.row_dimensions[r7].height = 18

    ws1.row_dimensions[ttd_row + 1].height = 18
    ws1.row_dimensions[ttd_row + 2].height = 20
    ws1.row_dimensions[ttd_row + 6].height = 22

    # Footer
    foot_row = ttd_row + 9
    ws1.merge_cells(f'A{foot_row}:I{foot_row}')
    c = ws1[f'A{foot_row}']
    c.value = (
        f'Dicetak: {date.today().strftime("%d %B %Y")}   |   '
        'SMA Negeri 5 Kota X - Sistem Informasi Akademik   |   '
        'Dokumen ini sah tanpa stempel basah'
    )
    c.font      = Font(name='Arial', size=8, color=WHITE, italic=True)
    c.fill      = _fill(NAVY)
    c.alignment = _align('center')
    ws1.row_dimensions[foot_row].height = 18

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — REKAP KEHADIRAN
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Rekap Kehadiran')
    ws2.sheet_properties.tabColor = '15803D'
    ws2.sheet_view.showGridLines  = False

    # FIX: lebar kolom diperbaiki agar Status Kehadiran tidak terpotong
    # A=No(6) | B=Status Kehadiran(20) | C=Jumlah Hari(14) | D=Persentase(14)
    # E-G=merge keterangan(14 each) | H=Keterangan(20)
    for col_ltr, w in [('A',18),('B',20),('C',14),('D',14),('E',14),('F',14),('G',14),('H',20)]:
        ws2.column_dimensions[col_ltr].width = w

    # Banner
    ws2.merge_cells('A1:H1')
    c = ws2['A1']
    c.value     = f'  \u258c  REKAP KEHADIRAN SISWA  -  {siswa.nama_lengkap}  ({nama_kelas})'
    c.font      = Font(name='Arial', bold=True, color=WHITE, size=12)
    c.fill      = _fill(NAVY)
    c.alignment = _align('center')
    ws2.row_dimensions[1].height = 32

    # Sub-header
    ws2.merge_cells('A2:H2')
    c = ws2['A2']
    c.value     = 'Tahun Akademik 2025/2026  |  Semester Ganjil  |  ' + date.today().strftime('%d %B %Y')
    c.font      = Font(name='Arial', size=9, color=GREY_MED, italic=True)
    c.fill      = _fill(NAVY_XLIGHT)
    c.alignment = _align('center')
    ws2.row_dimensions[2].height = 18

    # Spacer
    ws2.row_dimensions[3].height = 14

    # Identitas
    id2 = [
        ('Nama Lengkap',     siswa.nama_lengkap, NAVY,      True),
        ('Orang Tua / Wali', nama_wali,          GREY_DARK, False),
        ('Kelas',            nama_kelas,         GREY_DARK, True),
        ('Wali Kelas',       wali_kelas_nama,    GREY_DARK, False),
    ]
    for i, (lbl, val, vc, vb) in enumerate(id2, start=4):
        bg  = GREY_LIGHT if i % 2 else WHITE
        c_l = ws2.cell(row=i, column=1, value=lbl)
        c_l.font      = _font(bold=True, color=GREY_MED)
        c_l.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_l.fill      = _fill(bg)
        c_l.border    = _thin()
        c_s = ws2.cell(row=i, column=2, value=':')
        c_s.font      = _font(color=GREY_MED)
        c_s.fill      = _fill(bg)
        c_s.border    = _thin()
        c_s.alignment = _align('center')
        ws2.merge_cells(f'C{i}:H{i}')
        c_v = ws2.cell(row=i, column=3, value=val)
        c_v.font      = _font(bold=vb, color=vc, size=10 if vb else 9)
        c_v.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_v.fill      = _fill(bg)
        c_v.border    = _thin()
        for col in range(4, 9):
            ws2.cell(row=i, column=col).fill   = _fill(bg)
            ws2.cell(row=i, column=col).border = _thin()
        ws2.row_dimensions[i].height = 22

    # Spacer sebelum tabel
    ws2.row_dimensions[8].height = 14

    # Header tabel kehadiran
    ATT_ROW  = 9
    att_hdrs = ['No', 'Status Kehadiran', 'Jumlah Hari', 'Persentase', '', '', '', 'Keterangan']
    for col, h in enumerate(att_hdrs, start=1):
        c = ws2.cell(row=ATT_ROW, column=col, value=h)
        c.font      = Font(name='Arial', bold=True, color=WHITE, size=9)
        c.fill      = _fill(NAVY)
        c.alignment = _align('center')
        c.border    = _thin()
    ws2.row_dimensions[ATT_ROW].height = 24

    # Data kehadiran
    att_detail = [
        ('Hadir',  hadir_count, f'{pct_hadir}%',                                                 GREEN,  'Kehadiran aktif dalam KBM'),
        ('Izin',   izin_count,  f'{int(izin_count/total_absen*100) if total_absen else 0}%',      ORANGE, 'Tidak hadir dengan keterangan resmi'),
        ('Sakit',  sakit_count, f'{int(sakit_count/total_absen*100) if total_absen else 0}%',     ORANGE, 'Tidak hadir karena sakit'),
        ('Alpa',   alpa_count,  f'{int(alpa_count/total_absen*100) if total_absen else 0}%',      RED,    'Tidak hadir tanpa keterangan'),
        ('Total',  total_absen, '100%',                                                           NAVY,   'Total hari efektif semester ini'),
    ]
    for i, (status, jml, pct, clr, ket) in enumerate(att_detail):
        r   = ATT_ROW + 1 + i
        bg  = GREY_LIGHT if i % 2 else WHITE

        # No
        c = ws2.cell(row=r, column=1, value=i+1)
        c.font      = _font(color=GREY_MED)
        c.fill      = _fill(bg)
        c.border    = _thin()
        c.alignment = _align()

        # Status Kehadiran — kolom B, lebar sudah 20
        c = ws2.cell(row=r, column=2, value=status)
        c.font      = Font(name='Arial', bold=True, size=11, color=clr)
        c.fill      = _fill(bg)
        c.border    = _thin()
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        # Jumlah Hari
        c = ws2.cell(row=r, column=3, value=jml)
        c.font      = Font(name='Arial', bold=True, size=13, color=clr)
        c.fill      = _fill(bg)
        c.border    = _thin()
        c.alignment = _align()

        # Persentase
        c = ws2.cell(row=r, column=4, value=pct)
        c.font      = _font(size=9, color=GREY_MED)
        c.fill      = _fill(bg)
        c.border    = _thin()
        c.alignment = _align()

        # Keterangan — merge E:H penuh, tidak ada kolom bolong
        ws2.merge_cells(f'E{r}:H{r}')
        c = ws2.cell(row=r, column=5, value=ket)
        c.font      = _font(color=GREY_MED, italic=True)
        c.fill      = _fill(bg)
        c.border    = _thin()
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        # FIX: sel F,G,H dalam merge tetap dapat fill & border
        for col in range(6, 9):
            ws2.cell(row=r, column=col).fill   = _fill(bg)
            ws2.cell(row=r, column=col).border = _thin()

        ws2.row_dimensions[r].height = 26

    # Baris total ringkasan bawah — fill penuh A:H
    sum_row = ATT_ROW + len(att_detail) + 1
    ws2.row_dimensions[sum_row].height = 14 # spacer

    ring_row = sum_row + 1
    ws2.merge_cells(f'A{ring_row}:H{ring_row}')
    pct_color = GREEN if pct_hadir >= 75 else RED
    c = ws2.cell(row=ring_row, column=1,
                 value=f'Persentase Kehadiran : {pct_hadir}%   |   '
                       f'Hadir {hadir_count} hari dari {total_absen} hari efektif')
    c.font      = Font(name='Arial', bold=True, size=10, color=pct_color)
    c.fill      = _fill(GREEN_LIGHT if pct_hadir >= 75 else RED_LIGHT)
    c.alignment = _align('center')
    c.border    = _thin(NAVY)
    for col in range(2, 9):
        ws2.cell(row=ring_row, column=col).fill   = _fill(GREEN_LIGHT if pct_hadir >= 75 else RED_LIGHT)
        ws2.cell(row=ring_row, column=col).border = _thin(NAVY)
    ws2.row_dimensions[ring_row].height = 26

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — NILAI PERILAKU
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Nilai Perilaku')
    ws3.sheet_properties.tabColor = 'CA8A04'
    ws3.sheet_view.showGridLines  = False
    for col_ltr, w in [('A',24),('B',3),('C',35),('D',20)]:
        ws3.column_dimensions[col_ltr].width = w

    ws3.merge_cells('A1:D1')
    c = ws3['A1']
    c.value = f'  \u258c  PENILAIAN PERILAKU DAN SIKAP  -  {siswa.nama_lengkap}'
    c.font  = Font(name='Arial', bold=True, color=WHITE, size=12)
    c.fill  = _fill(NAVY)
    c.alignment = _align('center')
    ws3.row_dimensions[1].height = 32

    ws3.merge_cells('A2:D2')
    c = ws3['A2']
    c.value = 'Kontribusi terhadap Nilai Akhir : 15%   |   Skala : Baik = 100  Cukup = 70  Kurang = 30'
    c.font  = Font(name='Arial', size=9, color=GREY_MED, italic=True)
    c.fill  = _fill(GOLD_LIGHT)
    c.alignment = _align('center')
    ws3.row_dimensions[2].height = 18

    warna_prl = GREEN if nilai_perilaku_global >= 70 else ORANGE if nilai_perilaku_global >= 50 else RED

    prl_rows = [
        ('Nama Siswa',        siswa.nama_lengkap,                           NAVY,      True),
        ('Orang Tua / Wali',  nama_wali,                                    GREY_DARK, False),
        ('Skor Perilaku',     f'{nilai_perilaku_global} / 100',             warna_prl, True),
        ('Kategori',          _label_perilaku(nilai_perilaku_global),        warna_prl, True),
        ('Total Penilaian',   f'{total_penilaian} penilaian dari guru',     GREY_DARK, False),
        ('Distribusi Baik',   f'{perilaku_baik} penilaian (skor 100)',      GREEN,     False),
        ('Distribusi Cukup',  f'{perilaku_cukup} penilaian (skor 70)',      ORANGE,    False),
        ('Distribusi Kurang', f'{perilaku_kurang} penilaian (skor 30)',     RED,       False),
        ('Kontribusi Rapor',  '15% dari nilai akhir tiap mata pelajaran',   GREY_MED,  False),
    ]
    for i, (lbl, val, clr, bld) in enumerate(prl_rows, start=4):
        bg  = GREY_LIGHT if i % 2 else WHITE
        c_l = ws3.cell(row=i, column=1, value=lbl)
        c_l.font      = _font(bold=True, color=GREY_MED)
        c_l.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_l.border    = _thin()
        c_l.fill      = _fill(bg)
        c_s = ws3.cell(row=i, column=2, value=':')
        c_s.font      = _font(color=GREY_MED)
        c_s.border    = _thin()
        c_s.fill      = _fill(bg)
        c_s.alignment = _align('center')
        ws3.merge_cells(f'C{i}:D{i}')
        c_v = ws3.cell(row=i, column=3, value=val)
        c_v.font      = Font(name='Arial', bold=bld, size=10 if bld else 9, color=clr)
        c_v.border    = _thin()
        c_v.fill      = _fill(bg)
        c_v.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws3.cell(row=i, column=4).fill   = _fill(bg)
        ws3.cell(row=i, column=4).border = _thin()
        ws3.row_dimensions[i].height = 22

    # ── Print setup semua sheet ───────────────────────────────
    for ws in [ws1, ws2, ws3]:
        ws.page_setup.orientation  = 'portrait'
        ws.page_setup.paperSize    = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage    = True
        ws.page_setup.fitToWidth   = 1
        ws.page_setup.fitToHeight  = 0
        ws.page_margins.left       = 0.5
        ws.page_margins.right      = 0.5
        ws.page_margins.top        = 0.75
        ws.page_margins.bottom     = 0.75

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nama_file = f'Rapor_{siswa.nama_lengkap.replace(" ", "_")}_2025-2026.xlsx'
    response  = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nama_file}"'
    return response


# ============================================================
# VIEW 4 — Export Excel semua siswa
# PATCH: tambah kolom "Orang Tua / Wali" dan 'wali' di select_related
# ============================================================
@login_required
def export_semua_siswa_excel(request):
    user = request.user
    if not (user.is_staff or user.is_superuser or hasattr(user, 'kepsek_core_user')):
        raise Http404("Akses ditolak.")

    # PATCH: tambah 'wali' ke select_related
    semua_siswa = Siswa.objects.select_related(
        'kelas', 'kelas__wali_kelas', 'wali',
    ).order_by('kelas__nama_kelas', 'nama_lengkap')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data Siswa'

    NAVY_HEX  = '1E3A8A'
    NAVY2_HEX = '1E3A8A'
    GREY_HEX  = 'F8FAFC'
    WHITE_HEX = 'FFFFFF'
    GREEN_HEX = '15803D'
    RED_HEX   = 'B91C1C'

    def thin_border_all():
        s = Side(style='thin', color='D1D5DB')
        return Border(left=s, right=s, top=s, bottom=s)

    ws.merge_cells('A1:K1')   # PATCH: kolom bertambah 1 (K, dari J)
    ws['A1']           = 'REKAP DATA SISWA  —  SMA NEGERI 5 KOTA X'
    ws['A1'].font      = Font(bold=True, color=NAVY_HEX, size=13)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:K2')
    ws['A2']           = f'Dicetak: {date.today().strftime("%d %B %Y")}'
    ws['A2'].font      = Font(color='6B7280', size=9)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:K3')
    ws['A3'] = 'Formula Nilai Akhir: Tugas (35%) + UTS (25%) + UAS (25%) + Perilaku (15%)  |  KKM: 70'
    ws['A3'].font      = Font(italic=True, color='6B7280', size=8.5)
    ws['A3'].alignment = Alignment(horizontal='center')

    # PATCH: tambah kolom "Orang Tua / Wali" setelah "Wali Kelas"
    headers    = ['No', 'NISN', 'Nama Lengkap', 'Kelas', 'Wali Kelas',
                  'Orang Tua / Wali',           # ← kolom baru
                  'Rata-rata Nilai', 'Nilai Perilaku', '% Kehadiran',
                  'Status Keuangan', 'Total Mapel']
    col_widths = [5, 15, 30, 12, 25,
                  25,                            # ← lebar kolom baru
                  14, 14, 14, 18, 12]

    for col, (h, cw) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=5, column=col, value=h)
        cell.font      = Font(bold=True, color=WHITE_HEX, size=10)
        cell.fill      = PatternFill('solid', fgColor=NAVY2_HEX if col == 8 else NAVY_HEX)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border_all()
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = cw
    ws.row_dimensions[5].height = 28

    for i, siswa in enumerate(semua_siswa, start=1):
        row              = 5 + i
        nilai_qs         = NilaiRapor.objects.filter(siswa=siswa)
        total_mapel_s    = nilai_qs.count()
        nilai_perilaku_s = _hitung_nilai_perilaku(siswa)

        if total_mapel_s > 0:
            total_s = sum(
                _hitung_nilai_akhir(n.nilai_tugas, n.nilai_uts, n.nilai_uas, nilai_perilaku_s)
                for n in nilai_qs
            )
            rata_s = round(total_s / total_mapel_s, 1)
        else:
            rata_s = ''

        presensi_qs_s = Presensi.objects.filter(siswa=siswa)
        total_s2      = presensi_qs_s.count()
        hadir_s       = presensi_qs_s.filter(status='Hadir').count()
        pct_s         = f'{int(hadir_s / total_s2 * 100)}%' if total_s2 else ''

        wali_kelas_s = '-'
        if siswa.kelas and siswa.kelas.wali_kelas:
            wali_kelas_s = siswa.kelas.wali_kelas.nama_lengkap

        # PATCH: gunakan _nama_wali()
        nama_wali_s = _nama_wali(siswa)

        total_tagihan_s = KeuanganSiswa.objects.filter(siswa=siswa).count()
        tagihan_belum_s = KeuanganSiswa.objects.filter(siswa=siswa, status_lunas=False).count()

        if total_tagihan_s == 0:
            status_keu, warna_keu = 'Belum Ada Data', '6B7280'
        elif tagihan_belum_s > 0:
            status_keu, warna_keu = f'Tunggakan ({tagihan_belum_s})', RED_HEX
        else:
            status_keu, warna_keu = 'Lunas', GREEN_HEX

        warna_prl = (
            GREEN_HEX if nilai_perilaku_s >= 70
            else '92400E' if nilai_perilaku_s >= 50
            else RED_HEX
        )

        # PATCH: data_row sekarang 11 kolom (tambah nama_wali_s di index 5)
        data_row = [
            i,
            siswa.nisn,
            siswa.nama_lengkap,
            siswa.kelas.nama_kelas if siswa.kelas else '-',
            wali_kelas_s,
            nama_wali_s,       # ← PATCH
            rata_s,
            nilai_perilaku_s,
            pct_s,
            status_keu,
            total_mapel_s,
        ]

        for col, val in enumerate(data_row, start=1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.border    = thin_border_all()
            cell.alignment = Alignment(
                horizontal='left' if col in (3, 5, 6) else 'center',
                vertical='center'
            )
            if col == 8:    # Nilai Perilaku
                cell.font = Font(size=9, bold=True, color=warna_prl)
            elif col == 10: # Status Keuangan
                cell.font = Font(size=9, bold=True, color=warna_keu)
            elif col == 9:  # % Kehadiran
                cell.font = Font(size=9, bold=True, color=NAVY_HEX)
            else:
                cell.font = Font(size=9, color='111827')

            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=GREY_HEX)

        ws.row_dimensions[row].height = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Rekap_Siswa_2025-2026.xlsx"'
    return response